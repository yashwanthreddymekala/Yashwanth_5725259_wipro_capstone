import os
from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def read_excel(file_name, sheet_name):
        data = []

        base_dir = os.path.dirname(os.path.dirname(__file__))
        file_path = str(os.path.join(base_dir, "data", file_name))

        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)

        return data